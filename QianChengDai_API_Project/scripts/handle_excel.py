# -*- coding utf-8 -*-

# 该模块中实现对Excel文件的操作封装
from openpyxl import load_workbook
from scripts.handle_config import conf_operate


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        """
        将文件名、表单名作为实例属性
        :param filename:
        :param sheetname:
        """
        self.filename, self.sheetname = filename, sheetname

    def get_all_case(self):
        """
        该方法用于获取所有的测试数据，并返回一个嵌套字典的列表
        :return:
        """
        # 1. 创建文件对象
        wb = load_workbook(self.filename)
        # 2. 指定表单
        # 如果表单名为None则默认取第一个表单
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        # 3. 获取表头信息
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]  # 此处还可转换为列表等其他序列类型，但元组的效率更高

        # 4. 获取所有测试数据
        case_data = []
        # FileClass = namedtuple("FileClass", head_tuple)  # 创建一个类
        for item in tuple(ws.iter_rows(min_row=2, values_only=True)):
            # file_object = FileClass._make(item)  # 使用每行测试数据创建命名元组对象
            dict_data = dict(zip(head_tuple, item))
            case_data.append(dict_data)  # 将命名元组对象作为元素添加到列表中，就可通过属性名取值方式方便的取出想要的测试数据

        return case_data

    def get_one_case(self, row):
        """
        获取指定行的测试数据
        :return:
        """
        return self.get_all_case()[row-1]

    def write_data_in_excel(self, row, real_result, case_result, check_sql_result=None):
        real_result_column = conf_operate.get_int_config_data("result column", "real_result")
        case_result_column = conf_operate.get_int_config_data("result column", "case_result")

        # 注意：此处使用的Excel文件对象要与写入方法中的文件对象不同，否则会出现同一个workbook对象在对多个表单进行写入时只能将最后一个表单写入成功
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        # Excel文件中将实际结果、用例最终执行结果所在行号分别固定为6、7
        # 对传入行row参数进行判断，防止覆盖掉表头或超出最大行号
        if isinstance(row, int) and (2 <= row <= ws.max_row):
            ws.cell(row=row, column=real_result_column, value=real_result)
            ws.cell(row=row, column=case_result_column, value=case_result)
            if check_sql_result:
                ws.cell(row=row, column=10, value=check_sql_result)
        else:
            print("输入行号有误！")

        # 写入操作后，一定要保存文件
        wb.save(self.filename)


if __name__ == '__main__':
    pass
    # excel_operate = HandelExcel(r"D:\lyx_python\QianChengDai_API_Project\datas\testcase_data.xlsx")
    # print(excel_operate.get_all_case())
    # pass
    # print(excel_operate.get_one_case(3))
    # excel_operate.write_data_in_excel(2, "实际结果", "Pass")
