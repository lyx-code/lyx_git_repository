# -*- coding: utf-8 -*-
# @Time    : 2019/8/19 14:14
# @Author  : lyx
# @File    : handle_excel.py
# @Project : WebService_Api_Test
# 此模块为处理excel文件读取操作
import os
from openpyxl import load_workbook
from collections import namedtuple
from scripts.handle_config import HandleConfig
from scripts.content_os import CONFIG_FILE_PATH
do_conf = HandleConfig(os.path.join(CONFIG_FILE_PATH, "api_config.ini"))


class HandleExcel:
    """
    处理excel文件读取操作封装类
    """
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_all_case(self):
        """
        读取excel文件中所有用例数据
        :return:  返回嵌套用例数据字典的列表
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        cases = []
        header_data = tuple(ws.iter_rows(max_row=2, values_only=True))[0]
        CaseData = namedtuple("CaseData", header_data)
        data = tuple(ws.iter_rows(min_row=2, values_only=True))
        for item in data:
            case_data = CaseData._make(item)
            case = case_data._asdict()
            cases.append(dict(case))

        return cases

    def get_one_case(self, row):
        """
        读取一条用例数据
        :param row:
        :return:
        """
        return self.get_all_case()[row-1]

    def write_in_excel(self, row, actual, result, check_result=None):
        """
        往excel文件中写入数据
        :param row:
        :param actual:
        :param result:
        :param check_result:
        :return:
        """
        # 写入excel时，不要和上面读取时的对象使用同一个，否则会出现写入多个表单时只有最后一个表单能写入成功
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        ws.cell(row=row+1, column=do_conf.get_int_data("excel result", "actual"), value=actual)
        ws.cell(row=row+1, column=do_conf.get_int_data("excel result", "result"), value=result)
        ws.cell(row=row+1, column=do_conf.get_int_data("excel result", "check_result"), value=check_result)
        # excel文件写入后一定要save
        wb.save(self.filename)


if __name__ == '__main__':
    do = HandleExcel(r"D:\PythonLearning\WebService_Api_Test\datas\test_cases.xlsx", "login")
    cases = do.get_all_case()
    print(f"所有测试数据为：\n {cases}")
    case = do.get_one_case(4)
    print(f"第4行测试数据为：\n {case}")

    actual = "Webservice_test_actual"
    result = "Webservice_test_result"
    do.write_in_excel(5, actual, result)


